import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
import sys
import argparse
import re

def process_sarif(input_file_path):
    # Load SARIF file with utf-8 encoding
    with open(input_file_path, 'r', encoding='utf-8') as file:
        sarif = json.load(file)

    # Extract the results section
    results = sarif.get('runs', [])[0].get('results', [])
    
    # Severity mapping from SARIF levels to normalized levels
    severity_mapping = {
        'note': 'Low',
        'warning': 'Medium',
        'error': 'High'
    }

    # Process results into a DataFrame
    rows = []
    for result in results:
        file_path = result.get('locations', [{}])[0].get('physicalLocation', {}).get('artifactLocation', {}).get('uri', 'N/A')
        file_name = os.path.basename(file_path)  # Extract the file name from the path

        # Get the details text and replace Aquasec URLs with NIST URLs
        details_text = result.get('message', {}).get('text', 'N/A')
        details_text = details_text.replace('avd.aquasec.com/nvd/', 'nvd.nist.gov/vuln/detail/')

        # Normalize severity level
        raw_severity = result.get('level', 'N/A')
        normalized_severity = severity_mapping.get(raw_severity, raw_severity)

        rows.append({
            "Severity": normalized_severity,
            "Message": result.get('ruleId', 'N/A'),  # Renamed from Rule ID
            "Details": details_text,  # Renamed from Message
            "Path": file_path,  # Renamed from File
            "Page": file_name,  # Use the file name as the page value
            "Line": result.get('locations', [{}])[0].get('physicalLocation', {}).get('region', {}).get('startLine', 'N/A')
        })
    
    # Convert to DataFrame
    df = pd.DataFrame(rows, columns=["Severity", "Message", "Details", "Path", "Page", "Line"])

    # Generate output file path in same directory as input, with .xlsx extension
    input_dir = os.path.dirname(input_file_path)
    input_filename = os.path.basename(input_file_path)
    base_name = os.path.splitext(input_filename)[0]
    output_filename = base_name + '.xlsx'
    output_file = os.path.join(input_dir, output_filename)

    # Sanitize sheet name: remove Excel-disallowed characters and truncate to 31 chars
    sheet_name = base_name
    # Remove characters that Excel doesn't allow in sheet names: [ ] : \ / ? *
    for char in ['[', ']', ':', '\\', '/', '?', '*']:
        sheet_name = sheet_name.replace(char, '_')
    # Truncate to Excel's 31 character limit for sheet names
    sheet_name = sheet_name[:31]

    # Save to Excel
    df.to_excel(output_file, index=False, sheet_name=sheet_name)

    # Add table formatting and adjust column widths
    add_excel_table_and_adjust_columns(output_file, sheet_name, wrap_columns=["Message", "Details"], auto_fit_columns=["Path", "Page", "Line"])
    print(f"Processed Report saved to {output_file}")


def add_excel_table_and_adjust_columns(file_path, sheet_name, wrap_columns, auto_fit_columns):
    # Load the workbook and worksheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Determine the range of the table (e.g., A1:F20)
    start_cell = ws.cell(row=1, column=1)
    end_cell = ws.cell(row=ws.max_row, column=ws.max_column)
    table_range = f"{start_cell.coordinate}:{end_cell.coordinate}"

    # Create a table with filters but no banded rows
    table = Table(displayName="SARIFTable", ref=table_range)

    # Add a table style without banded rows
    style = TableStyleInfo(
        name="TableStyleLight1",  # Use a plain style
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,  # Disable banded rows
        showColumnStripes=False  # Disable banded columns
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet (this enables filters)
    ws.add_table(table)

    # Define styling elements
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='595959', end_color='595959', fill_type='solid')  # Dark grey (approximately 22% grey)

    data_font = Font(name='Calibri', size=11, bold=False, color='000000')
    data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White background

    # Apply styling to all populated cells
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Apply borders to all cells
            cell.border = thin_border

            # Apply font to all cells
            if row_idx == 1:
                # Header row styling
                cell.font = header_font
                cell.fill = header_fill
            else:
                # Data row styling
                cell.font = data_font
                cell.fill = data_fill

    # Adjust column widths and wrap text
    details_col_index = None
    for col_name in ws[1]:
        if col_name.value == "Details":
            details_col_index = col_name.column

        if col_name.value in auto_fit_columns:
            # Auto-fit column widths based on content length
            col_index = col_name.column
            column_letter = ws.cell(row=1, column=col_index).column_letter
            max_length = max(len(str(ws.cell(row=row, column=col_index).value or "")) for row in range(2, ws.max_row + 1))  # Exclude header
            ws.column_dimensions[column_letter].width = max_length + 2  # Add padding
        elif col_name.value in wrap_columns:
            # Adjust width to 3/4 of content and enable text wrapping
            col_index = col_name.column
            column_letter = ws.cell(row=1, column=col_index).column_letter
            max_length = max(len(str(ws.cell(row=row, column=col_index).value or "")) for row in range(2, ws.max_row + 1))  # Exclude header
            ws.column_dimensions[column_letter].width = (max_length * 0.75)  # Set width to 3/4 of content length
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_index)
                cell.alignment = Alignment(wrap_text=True)  # Enable text wrapping

    # Convert Markdown links to Excel hyperlinks in Details column
    if details_col_index:
        # Regex pattern to match Markdown links: [text](url)
        markdown_link_pattern = r'\[([^\]]+)\]\(([^\)]+)\)'

        for row_idx in range(2, ws.max_row + 1):  # Skip header row
            cell = ws.cell(row=row_idx, column=details_col_index)
            cell_value = str(cell.value or "")

            # Check if cell contains a Markdown link
            match = re.search(markdown_link_pattern, cell_value)
            if match:
                link_text = match.group(1)  # The text inside [...]
                link_url = match.group(2)   # The URL inside (...)

                # Replace the Markdown link with just the link text
                new_cell_value = re.sub(markdown_link_pattern, link_text, cell_value)

                # Update cell value
                cell.value = new_cell_value

                # Add hyperlink to the cell
                cell.hyperlink = link_url

                # Style the hyperlink (blue and underlined)
                cell.font = Font(name='Calibri', size=11, color='0563C1', underline='single')

    # Save the workbook
    wb.save(file_path)
    print(f"Table formatting and column adjustments added to {file_path}")

def main():
    parser = argparse.ArgumentParser(description='Convert SARIF reports to Excel format')
    parser.add_argument('sarif_file', help='Path to the SARIF file to process')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.sarif_file):
        print(f"Error: File '{args.sarif_file}' not found.")
        sys.exit(1)
    
    process_sarif(args.sarif_file)

if __name__ == '__main__':
    main()
